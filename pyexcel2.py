from openpyxl import load_workbook
workbook = load_workbook(filename = "reviews-sample.xlsx")
workbook.sheetnames

sheet = workbook.active

print(sheet.title)

print(sheet["F29"])

print(sheet["F29"].value)

print(sheet["a1:c2"])


#print(content_string)
